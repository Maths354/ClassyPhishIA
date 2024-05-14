import openpyxl
import logging

logging.basicConfig(level=logging.INFO)

def is_id_present(data_id, sheet):
    """Vérifie si l'ID existe déjà dans la feuille et retourne l'index de ligne si trouvé."""
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        existing_id = row[0]
        if existing_id == data_id:
            return index
    return None

def save_to_excel(data, filename):
    """Sauvegarder les données dans un fichier Excel."""
    data_id, url_legitime, sha256, parsed_tags = data
    
    # Charger ou créer le workbook et obtenir la feuille active
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Ajouter les en-têtes si c'est un nouveau fichier
        sheet.append(['ID', 'URL', 'SHA-256', 'BALISES'])

    # Vérifier si l'ID existe déjà et obtenir la ligne correspondante
    row_index = None
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        existing_id = row[0]
        if existing_id == data_id:
            row_index = index
            break

    if row_index is not None:
        # Mettre à jour les colonnes existantes
        sheet.cell(row=row_index, column=2, value=url_legitime)  # URL
        sheet.cell(row=row_index, column=3, value=sha256)  # SHA-256
        
        # Mettre à jour `parsed_tags` uniquement si une valeur valide est fournie
        if parsed_tags:
            sheet.cell(row=row_index, column=4, value=parsed_tags)  # BALISES
        # if urls_link:
        #     sheet.cell(row=row_index, column=5, value=urls_link)  # FUTURE_URL
        
        logging.info(f"Les données avec l'ID {data_id} ont été mises à jour.")
    else:
        # Ajouter les données à la feuille
        sheet.append(data)
        logging.info(f"Les données ont été ajoutées au fichier Excel : {filename}")
    
    # Enregistrer le fichier Excel
    workbook.save(filename)
    workbook.close()

